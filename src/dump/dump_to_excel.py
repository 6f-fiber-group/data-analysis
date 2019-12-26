import os
import sys
import yaml
import datetime
import openpyxl

path = os.path.join(os.path.dirname(__file__), "../features")
sys.path.append(path)

from old_spectrum_analyzer_manager import OldSpectrumAnalyzerManager as OSAM
from new_spectrum_analyzer_manager import NewSpectrumAnalyzerManager as NSAM

def dump(data, params):
  wb = openpyxl.Workbook()
  sheet = wb.active

  sheet.cell(1,1, "x")
  for col_idx_x, d in enumerate(data["x"].items()):
    label, x = d
    sheet.cell(2, col_idx_x + 1, label)
    for row_idx, val in enumerate(x):
      sheet.cell(row_idx+3, col_idx_x + 1, val)

  sheet.cell(1, col_idx_x + 3, "y")
  for col_idx_y, d in enumerate(data["y"].items()):
    label, y = d
    sheet.cell(2, col_idx_x + col_idx_y + 3, label)
    for row_idx, val in enumerate(y):
      sheet.cell(row_idx+3, col_idx_x + col_idx_y + 3, val)

  save(wb)

def save(wb):
  save_file_path = os.path.join(params["data_path"], "excel")

  if not os.path.exists(save_file_path):
    os.makedirs(save_file_path)
  if not params["file_name"]:
    params["file_name"] = str(datetime.date.today())

  wb.save(os.path.join(save_file_path, "{}.xlsx".format(params["file_name"])))

def get_dump_data(params):
  if params["equip_type"] == "old_spectrum_analyzer":
    sobj = OSAM(params["data_path"])
  elif params["equip_type"] == "new_spectrum_analyzer":
    sobj = NSAM(params["data_path"])

  if params["calc_type"] == "diff":
    return sobj.diff(params["std_data_file"], params["convolve"])
  elif params["calc_type"] == "normalize":
    return sobj.normalize(params["std_data_file"], params["convolve"])
  else:
    return sobj.raw_data(params["convolve"])

if __name__ == "__main__":
  data_path = r"put path here"
  with open(os.path.join(data_path, "params.yml"), "r") as f:
    params = yaml.safe_load(f)
    params["data_path"] = data_path
  
  data = get_dump_data(params)
  dump(data, params)